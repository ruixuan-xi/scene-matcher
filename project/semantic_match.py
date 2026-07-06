# ─────────────────────────────────────────────────────────────
# scene_matcher — 语义向量场景匹配
# 架构：Supabase REST API + sentence-transformers + numpy
# ─────────────────────────────────────────────────────────────

import os, sys, time, argparse, hashlib, re, ast, json, uuid
from typing import Optional

import numpy as np
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer

load_dotenv()

# ═══════════════════════════════════════════════════════
# 全局常量
# ═══════════════════════════════════════════════════════

REQUIRED_HEADERS = [
    "部门",
    "需求提交人",
    "工作流程名称",
    "具体工作步骤说明",
    "业务痛点",
    "涉及到的软件系统和网页",
    "企业名称",
]

STD_DEPT_REFERENCES = [
    "财务", "采购", "人事", "行政", "市场", "技术",
    "运营", "客服", "物流", "法务", "合规", "审计",
    "战略", "投融资", "董办", "综合",
]

# ── 步骤编号正则（统一常量 + 预编译）──
STEP_MARKER_PATTERN = (
    r"(?:\d{1,2}[\.\、，,]|[一二三四五六七八九十]{1,2}[\.\、，,]|[①②③④⑤⑥⑦⑧⑨⑩]|步骤\s*\d|第[一二三四五六七八九十\d]{1,2}步)"
)
_STEP_MARKER_PAT = re.compile(STEP_MARKER_PATTERN)

# ── Excel 样式常量（统一管理，避免重复定义）──
_STYLE_HEADER_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_STYLE_HEADER_FONT   = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_STYLE_YELLOW_FILL   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_STYLE_WARN_FONT     = Font(name="微软雅黑", size=10, color="C00000", bold=True)
_STYLE_THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_STYLE_WRAP          = Alignment(wrap_text=True, vertical="top")
_STYLE_GREEN_FILL    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_STYLE_BLUE_FILL     = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_STYLE_DEPT_FILL     = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
_STYLE_DEPT_FONT     = Font(name="微软雅黑", size=10, color="38761D")
_STYLE_DEPT_HFONT    = Font(name="微软雅黑", bold=True, size=11, color="38761D")
_STYLE_FEEDBACK_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
_STYLE_FEEDBACK_FONT = Font(name="微软雅黑", size=10, italic=True, color="C00000")
_STYLE_YELLOW2_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_STYLE_RED_FILL      = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_STYLE_CELL_FONT     = Font(name="微软雅黑", size=10)

# LLM 调用进程内缓存
_llm_cache = {}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "需求收集模板.xlsx")
MODEL_CACHE_DIR = os.path.join(BASE_DIR, ".model_cache")
HIGH_THRESHOLD = 0.85
LOW_THRESHOLD = 0.75
EMBEDDING_DIM = 384

os.environ.setdefault("HF_ENDPOINT", "https://hf-mirror.com")


# ═══════════════════════════════════════════════════════
# LLM 调用
# ═══════════════════════════════════════════════════════

def _llm_call(system_prompt: str, user_prompt: str) -> Optional[str]:
    """调用 DeepSeek API，失败时打印具体错误而非静默吞掉"""
    api_key = os.getenv("DEEPSEEK_API_KEY", "")
    if not api_key:
        print("⚠️ 未设置 DEEPSEEK_API_KEY")
        return None
    cache_key = hashlib.md5((system_prompt + user_prompt).encode()).hexdigest()
    if cache_key in _llm_cache:
        return _llm_cache[cache_key]
    try:
        r = requests.post(
            os.getenv("DEEPSEEK_API_URL", "https://api.deepseek.com/v1/chat/completions"),
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": "deepseek-chat",
                "temperature": 0,
                "max_tokens": 200,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
            },
            timeout=15,
        )
        if r.status_code == 200:
            result = r.json()["choices"][0]["message"]["content"].strip()
            _llm_cache[cache_key] = result
            return result
        elif r.status_code == 401:
            print(f"❌ DeepSeek API 认证失败（401）：请检查 DEEPSEEK_API_KEY")
        elif r.status_code == 429:
            print(f"⚠️ DeepSeek API 限流（429）：请稍后重试")
        else:
            print(f"❌ DeepSeek API 错误 {r.status_code}：{r.text[:200]}")
    except requests.exceptions.Timeout:
        print("❌ DeepSeek API 请求超时（15s）")
    except requests.exceptions.ConnectionError as e:
        print(f"❌ DeepSeek API 连接失败：{e}")
    except Exception as e:
        print(f"❌ DeepSeek API 未知错误：{e}")
    return None


# ═══════════════════════════════════════════════════════
# 部门归一化（批量 LLM）
# ═══════════════════════════════════════════════════════

def normalize_depts_batch(raw_depts: list[str]) -> dict[str, str]:
    """一次 LLM 调用批量归一化所有唯一部门名，返回 {原名: 归一化名}"""
    unique = list(dict.fromkeys(d.strip() for d in raw_depts if d and d.strip()))
    if not unique:
        return {}
    prompt = (
        "你是部门名称归一化助手。将以下部门名称归一化为简洁的标准名称（去掉「部」「中心」「科」等后缀）。\n"
        f"参考分类：{'、'.join(STD_DEPT_REFERENCES)}\n"
        "输出 JSON 格式：{\"原部门名\": \"归一化名称\", ...}。\n"
        "找不到合适分类则归一化为原文。只输出 JSON，不要解释。"
    )
    result = _llm_call(prompt, json.dumps(unique, ensure_ascii=False))
    if result:
        try:
            mapping: dict = json.loads(result)
            return {k: v for k, v in mapping.items() if k in unique}
        except Exception:
            pass
    return {d: d for d in unique}


def normalize_dept(raw: str) -> str:
    """单个部门归一化（内部委托批量接口）"""
    if not raw:
        return ""
    key = raw.strip()
    mapping = normalize_depts_batch([key])
    return mapping.get(key, key)


# ═══════════════════════════════════════════════════════
# 表头映射 & 列定位
# ═══════════════════════════════════════════════════════

def map_headers(input_headers: list[str]) -> tuple[dict, list]:
    """LLM 语义映射表头 → 标准列，返回 (mapping, unmapped)"""
    headers = [str(h).strip() if h else "" for h in input_headers]
    result = _llm_call(
        "你是Excel表头标准化助手。将输入表头映射到标准表头，输出JSON：{\"原列名\": \"标准列名\"}。找不到的值填 null。",
        f"输入表头：{json.dumps(headers, ensure_ascii=False)}\n"
        f"标准表头：{json.dumps(REQUIRED_HEADERS, ensure_ascii=False)}",
    )
    mapping, used, unmapped = {}, set(), []
    if result:
        try:
            llm_map: dict = json.loads(result)
            for req in REQUIRED_HEADERS:
                idx = None
                for i, ih in enumerate(headers):
                    if not ih or i in used:
                        continue
                    lv = llm_map.get(ih, "")
                    # 双向包含匹配（处理 "流程名" ↔ "流程名称" 差异）
                    if lv == req or req in lv or lv in req:
                        idx = i
                        break
                mapping[req] = idx
                if idx is not None:
                    used.add(idx)
                else:
                    unmapped.append(req)
            return mapping, unmapped
        except Exception:
            pass
    for req in REQUIRED_HEADERS:
        mapping[req] = None
        unmapped.append(req)
    return mapping, unmapped


def find_cols(headers: list) -> tuple[int, int, int, int, int, int, int]:
    """定位 Excel 各列位置。
    策略：先跑启发式（始终返回有效 int）→ LLM 语义定位（找到则覆盖）。
    保证返回值始终为 int，避免 ws.cell(r, None+1) 崩溃。
    """
    input_headers = [str(h).strip() if h else "" for h in headers]

    # ── 1. 启发式兜底（始终返回 int）──
    tc = sc = dc = oc = src = pl = pc = 0
    for i, h in enumerate(input_headers):
        s = h.strip()
        if (
            ("流程名称" in s or "工作流程" in s or "场景名称" in s or "流程标题" in s
             or (("场景" in s or "标题" in s) and "流程" not in s and "是否" not in s and "Top" not in s))
            and "步骤" not in s and "企业" not in s
        ):
            tc = i
        elif "步骤" in s or "说明" in s:
            sc = i
        elif "部门" in s:
            dc = i
        elif "提交人" in s or "负责" in s:
            oc = i
        elif "企业" in s or "公司" in s or "来源" in s:
            src = i
        elif "平台" in s or "系统" in s or "软件" in s:
            pl = i
        elif "痛点" in s or "耗时" in s or "频率" in s or "问题" in s:
            pc = i

    # ── 2. LLM 语义定位（找到非 null 值则覆盖）──
    result = _llm_call(
        "你是Excel列定位助手。给定表头，找出每类数据在第几列（从0开始计数）。\n"
        "七类数据：标题/流程名称、步骤/说明、部门、提交人/负责人、企业/来源、平台/系统/软件、痛点/问题\n"
        "输出 JSON 格式：{\"title\":0,\"steps\":1,\"dept\":2,\"owner\":3,\"source\":4,\"platform\":5,\"pain\":6}\n"
        "找不到则填 null。",
        f"表头：{json.dumps(input_headers, ensure_ascii=False)}",
    )
    if result:
        try:
            m: dict = json.loads(result)
            if isinstance(m.get("title"), int):    tc  = m["title"]
            if isinstance(m.get("steps"), int):    sc  = m["steps"]
            if isinstance(m.get("dept"), int):     dc  = m["dept"]
            if isinstance(m.get("owner"), int):    oc  = m["owner"]
            if isinstance(m.get("source"), int):   src = m["source"]
            if isinstance(m.get("platform"), int):  pl  = m["platform"]
            if isinstance(m.get("pain"), int):     pc  = m["pain"]
        except Exception:
            pass

    return tc, sc, dc, oc, src, pl, pc


# ═══════════════════════════════════════════════════════
# 步骤详细度检查
# ═══════════════════════════════════════════════════════

def check_steps_detail(text: str) -> tuple[bool, str, str]:
    """检查步骤是否足够详细（使用预编译正则）"""
    if not text or len(text.strip()) < 15:
        return False, "steps_too_short", "步骤说明过短或为空"

    # 宽松匹配（行首/句号/双空格 + 编号）
    loose = re.compile(r"(?:^|\n|。|；|\s{2,})" + STEP_MARKER_PATTERN)
    nidx = [m.start() for m in loose.finditer(text)]
    # 宽松模式只匹配到 <2 个 → 紧凑模式（如 "1、登录 2、打开"）
    if len(nidx) < 2:
        nidx = [m.start() for m in _STEP_MARKER_PAT.finditer(text)]
    if len(nidx) < 2:
        return False, "no_step_numbers", "缺少分步编号（如 1. 2. 3.），请按格式重写"

    # 去掉编号前缀后的正文长度
    body = _STEP_MARKER_PAT.sub("", text).strip()
    if len(body) < 30:
        return False, "steps_too_short", f"步骤总字数不足（当前 {len(body)} 字，建议 ≥30 字）"

    return True, "ok", "ok"


def _count_step_markers(text: str) -> int:
    """统计步骤编号标记数量（使用预编译正则）"""
    if not text:
        return 0
    return len(_STEP_MARKER_PAT.findall(text))


def validate_template(headers: list) -> tuple[bool, list]:
    """校验表头是否包含所有必填列（精确匹配，避免 "部门" 误匹配 "非部门"）"""
    header_texts = [str(h).strip() if h else "" for h in headers]
    missing = [h for h in REQUIRED_HEADERS if h not in header_texts]
    return len(missing) == 0, missing


# ═══════════════════════════════════════════════════════
# ID 生成 & DB 初始化
# ═══════════════════════════════════════════════════════

def _new_id() -> str:
    """生成 8 位十六进制 ID（uuid，无碰撞，无需 sleep）"""
    return uuid.uuid4().hex[:8]


def init_db(client: "SupabaseClient") -> None:
    """确保 scenes 表存在，插入一条样例场景"""
    try:
        client._post(
            "scenes",
            params={"on_conflict": "scene_id"},
            json={
                "scene_id": "fin_bank",
                "scene_title": "中国银行流水自动下载",
                "process_steps": "1.登录中国银行网银 2.进入企业账务管理-交易流水查询 3.设置日期范围（本月）并提交 4.导出Excel格式流水 5.保存至本地文件夹",
                "department": "财务",
                "platform": "中国银行网银",
                "owner": "张三",
                "source": "财务部",
                "pain": "每月手动登录下载流水，重复操作耗时约2小时",
            },
        )
        client._patch(
            "scenes",
            params={"scene_id": "eq.fin_bank"},
            json={"embedding": [0.05] * 384},
        )
    except Exception:
        pass


# ═══════════════════════════════════════════════════════
# Supabase 客户端（REST API 模式，无 psycopg2）
# ═══════════════════════════════════════════════════════

class SupabaseClient:
    def __init__(self):
        self.url = os.getenv("SUPABASE_URL", "").rstrip("/")
        self.key = os.getenv("SUPABASE_SERVICE_KEY", "")
        self._scenes: Optional[list] = None

    def _headers(self) -> dict:
        return {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

    def _get(self, table: str, params: Optional[dict] = None) -> list:
        r = requests.get(
            f"{self.url}/rest/v1/{table}",
            headers=self._headers(),
            params=params,
            timeout=20,
        )
        r.raise_for_status()
        return r.json() or []

    def _post(self, table: str, json: dict, params: Optional[dict] = None) -> requests.Response:
        p = dict(params) if params else {}
        r = requests.post(
            f"{self.url}/rest/v1/{table}",
            headers={
                **self._headers(),
                "Prefer": "return=representation,resolution=merge-duplicates",
            },
            params=p,
            json=json,
            timeout=20,
        )
        r.raise_for_status()
        return r

    def _patch(self, table: str, params: dict, json: dict) -> requests.Response:
        r = requests.patch(
            f"{self.url}/rest/v1/{table}",
            headers=self._headers(),
            params=params,
            json=json,
            timeout=20,
        )
        r.raise_for_status()
        return r

    def load_scenes(self) -> list:
        """拉取全量场景（含 embedding），缓存在本地。
        网络异常时以空基线继续（所有输入均为新场景）。
        """
        if self._scenes is not None:
            return self._scenes
        try:
            rows = self._get("scenes", params={
                "select": "id,scene_id,scene_title,process_steps,department,platform,owner,source,pain,embedding",
                "order": "id.asc",
            })
            for r in rows:
                emb = r.get("embedding")
                if emb is not None:
                    if isinstance(emb, str):
                        r["_vec"] = np.array(ast.literal_eval(emb), dtype=np.float32)
                    else:
                        r["_vec"] = np.array(emb, dtype=np.float32)
                else:
                    r["_vec"] = None
            self._scenes = rows
            return rows
        except Exception as e:
            print(f"⚠️ 加载场景库失败（将使用空基线）：{e}")
            self._scenes = []
            return []

    def upsert_scene(self, scene_data: dict) -> None:
        """插入/更新场景，清除本地缓存"""
        self._post("scenes", json=scene_data)
        self._scenes = None

    def match(self, query_vec, dept_std: Optional[str] = None, n: int = 1) -> list:
        """本地余弦相似度匹配（numpy 向量化运算），返回 Top-N (scene, sim)"""
        scenes = self.load_scenes()
        query_vec = np.array(query_vec, dtype=np.float32)
        q_norm = np.linalg.norm(query_vec)
        if q_norm == 0:
            return []

        # 部门硬过滤 + 提取向量
        filtered = [
            (s, np.array(s["_vec"], dtype=np.float32))
            for s in scenes
            if s["_vec"] is not None
            and (dept_std is None or (s.get("department") or "").strip() == dept_std)
        ]
        if not filtered:
            return []

        scene_vecs = np.stack([v for _, v in filtered])
        s_norms = np.linalg.norm(scene_vecs, axis=1)
        valid = s_norms > 0
        sims = np.zeros(len(filtered), dtype=np.float32)
        if valid.any():
            sims[valid] = np.dot(scene_vecs[valid], query_vec) / (s_norms[valid] * q_norm)

        order = np.argsort(-sims)
        return [(filtered[i][0], float(sims[i])) for i in order[:n]]

    def count(self) -> int:
        return len(self.load_scenes())


# ═══════════════════════════════════════════════════════
# 向量生成
# ═══════════════════════════════════════════════════════

def get_embedding(text: str) -> list:
    """调用 sentence-transformers 生成 embedding 向量"""
    # 优先使用 ModelScope 下载的本地模型（避免 HF Hub 连接问题）
    modelscope_path = os.path.join(
        MODEL_CACHE_DIR, "sentence-transformers", "all-MiniLM-L6-v2"
    )
    if os.path.isdir(modelscope_path):
        model_path = modelscope_path
    else:
        model_path = "sentence-transformers/all-MiniLM-L6-v2"
    model = SentenceTransformer(
        model_path,
        cache_folder=MODEL_CACHE_DIR,
    )
    vec = model.encode(text, normalize_embeddings=True)
    return vec.tolist()


def import_scene(
    client: SupabaseClient,
    scene_id: str,
    title: str,
    steps: str,
    dept: str,
    owner: str,
    platform: str,
    source: str,
    pain: str,
    dept_already_normalized: bool = False,
) -> None:
    """将场景写入 Supabase（含 embedding）"""
    dept_val = dept if dept_already_normalized else normalize_dept(dept)
    scene_data = {
        "scene_id": scene_id,
        "scene_title": title,
        "process_steps": steps,
        "department": dept_val,
        "owner": owner,
        "platform": platform,
        "source": source,
        "pain": pain,
    }
    try:
        client.upsert_scene(scene_data)
        print(f"  ✅ 已入库：{scene_id} | {title}")
        emb = get_embedding(f"{title} {steps} {pain}")
        client._patch(
            "scenes",
            params={"scene_id": f"eq.{scene_id}"},
            json={"embedding": emb},
        )
    except Exception as e:
        print(f"  ❌ 入库失败 {scene_id}：{e}")


def top_match(
    client: SupabaseClient,
    query_text: str,
    dept_std: Optional[str] = None,
    n: int = 1,
) -> tuple[Optional[dict], float]:
    """向量匹配，返回 (最佳匹配场景, 相似度)"""
    q_emb = get_embedding(query_text)
    results = client.match(q_emb, dept_std=dept_std, n=n)
    return results[0] if results else (None, 0.0)


# ═══════════════════════════════════════════════════════
# Excel 辅助函数
# ═══════════════════════════════════════════════════════

def _set_header(ws, row: int, col: int, value: str, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:  cell.fill = fill
    if font:  cell.font = font
    cell.border = _STYLE_THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _set_cell(ws, row: int, col: int, value, fill=None, font=None, wrap=True, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:  cell.fill = fill
    if font:  cell.font = font
    cell.border = _STYLE_THIN_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    return cell


def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _create_supplement_xlsx(unmapped: list, input_path: str) -> None:
    """生成补填模板 Excel（提示用户填写缺失的列）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "请补充以下列"

    ws.row_dimensions[1].height = 30
    _set_header(ws, 1, 1, "字段", _STYLE_HEADER_FILL, _STYLE_HEADER_FONT)
    _set_header(ws, 1, 2, "说明", _STYLE_HEADER_FILL, _STYLE_HEADER_FONT)
    _set_header(ws, 1, 3, "示例", _STYLE_HEADER_FILL, _STYLE_HEADER_FONT)

    tips = {
        "部门": '如"财务部"、"采购部"、"市场部"等',
        "需求提交人": "负责该需求的同事姓名",
        "工作流程名称": "简洁描述该工作流程的名称",
        "具体工作步骤说明": "分步描述操作步骤，每步用编号开头（如1. 2. 3.）",
        "业务痛点": "当前流程存在哪些问题或耗时点",
        "涉及到的软件系统和网页": "使用的系统或平台名称",
        "企业名称": "所属公司或部门名称",
    }

    for r, col_name in enumerate(unmapped, 2):
        _set_cell(ws, r, 1, col_name, _STYLE_YELLOW_FILL, _STYLE_WARN_FONT, align="center")
        _set_cell(ws, r, 2, tips.get(col_name, "请填写"))
        _set_cell(ws, r, 3, "示例内容", _STYLE_FEEDBACK_FILL, _STYLE_FEEDBACK_FONT, align="center")

    _col_widths(ws, [22, 35, 20])
    out_path = input_path.replace(".xlsx", "_需补充列.xlsx")
    wb.save(out_path)
    print(f"📋 请补充缺失列：{out_path}")


def _create_template() -> None:
    """生成需求收集模板 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "需求收集表"

    _col_widths(ws, [14, 12, 20, 45, 30, 25, 15])

    ws.row_dimensions[1].height = 35
    for i, h in enumerate(REQUIRED_HEADERS, 1):
        _set_header(ws, 1, i, h, _STYLE_HEADER_FILL, _STYLE_HEADER_FONT)

    title_font = Font(name="微软雅黑", size=11, bold=True, color="4472C4")
    tip_font   = Font(name="微软雅黑", size=9, color="666666")
    warn_font  = Font(name="微软雅黑", size=9, color="CC0000", bold=True)
    gray_font  = Font(name="微软雅黑", size=9, color="999999")

    instructions = [
        (4, "填写说明", title_font),
        (5, "① 所有7列均为必填项，缺一不可。如某列确实不适用，请填写「无」。", tip_font),
        (6, "②「具体工作步骤说明」必须用分步编号书写（1. 2. 3.），否则跳过匹配并以红色标注：", warn_font),
        (7, "✓ 正确：「1.登录京准通后台 → 2.点击数据报表-日报导出近7天数据 → 3.打开公司日报模板填入销售额、订单数 → 4.上传到共享文件夹」", tip_font),
        (8, "✓ 最少 30 字，每步写清楚在哪操作、操作什么", tip_font),
        (9, "✗ 反例1：「打开系统填表」→ 过短、无编号，会被退回", gray_font),
        (10, "✗ 反例2：「1、登录 2、打开网页 3、写入excel」→ 有编号但每步内容过于简略，也会被退回", gray_font),
        (11, "③ 请删除上方示例行后填入真实需求，每个需求占一行。", tip_font),
    ]

    for row_num, text, font in instructions:
        ws.row_dimensions[row_num].height = 18
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = font
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)

    ws.row_dimensions[12].height = 60
    examples = [
        "财务部", "李四", "中国银行流水下载",
        "1.登录中国银行网银 2.进入企业账务管理-交易流水查询 3.设置日期范围并提交 4.导出Excel格式流水 5.保存至本地文件夹",
        "每月手动登录下载流水，重复操作耗时约2小时", "中国银行网银", "示例公司A",
    ]
    for c, val in enumerate(examples, 1):
        cell = ws.cell(row=12, column=c, value=val)
        cell.fill = _STYLE_BLUE_FILL
        cell.font = _STYLE_CELL_FONT
        cell.border = _STYLE_THIN_BORDER
        cell.alignment = _STYLE_WRAP

    wb.save(TEMPLATE_PATH)
    print(f"📋 需求收集模板已生成：{TEMPLATE_PATH}")


def process_feedback(input_path: str) -> None:
    """处理用户填写的「是否入库」反馈列"""
    from openpyxl import load_workbook

    wb = load_workbook(input_path)
    ws = wb.active
    maxc = ws.max_column
    tc, sc, dc, oc, src, pl, pc = find_cols([ws.cell(1, c).value for c in range(1, maxc + 1)])

    client = SupabaseClient()
    imported = skipped = 0

    for r in range(2, ws.max_row + 1):
        fb_raw = ws.cell(r, maxc).value
        if not fb_raw or str(fb_raw).strip() not in ("是", "1", "Y", "y", "true", "是 "):
            continue
        t = ws.cell(r, tc + 1).value if tc < maxc else ""
        s = ws.cell(r, sc + 1).value if sc < maxc else ""
        # 部门已归一化（输出时回写过），直接用
        d = str(ws.cell(r, dc + 1).value).strip() if dc < maxc and dc is not None else ""
        o = ws.cell(r, oc + 1).value if oc < maxc and oc is not None else ""
        z = ws.cell(r, pl + 1).value if pl < maxc and pl is not None else ""
        pain = ws.cell(r, pc + 1).value if pc < maxc and pc is not None else ""

        if not t or not s:
            print(f"  ⏭ 跳过第{r}行（标题或步骤为空）")
            skipped += 1
            continue

        sid = _new_id()
        import_scene(client, sid, t, s, d, o, z, "", pain, dept_already_normalized=True)
        imported += 1

    print(f"\n✅ 反馈入库完成：{imported} 条已入库，{skipped} 条跳过")


# ═══════════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(description="语义向量场景匹配")
    parser.add_argument("--input", required=True, help="输入 Excel 路径")
    parser.add_argument("--auto-import", action="store_true", help="新场景自动入库")
    parser.add_argument("--apply-feedback", action="store_true", help="处理反馈入库")
    args = parser.parse_args()

    if args.apply_feedback:
        process_feedback(args.input)
        return

    input_path = args.input

    # 空路径 → 生成模板
    if input_path in ("/dev/null", "null", "none", ""):
        _create_template()
        return

    from openpyxl import load_workbook

    # ── Step 0: 表头校验 ──
    wb = load_workbook(input_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    ok, missing = validate_template(headers)
    if not ok:
        print(f"❌ 表头缺少必需列：{missing}")
        _create_supplement_xlsx(missing, input_path)
        return

    # 表头映射
    mapping, unmapped = map_headers(headers)
    if unmapped:
        print(f"⚠️ 无法自动映射的列：{unmapped}，将生成补填模板")
        _create_supplement_xlsx(unmapped, input_path)

    tc, sc, dc, oc, src, pl, pc = find_cols(headers)
    maxc = ws.max_column

    # ── Step 1-2: 连接场景库 ──
    client = SupabaseClient()
    init_db(client)
    total = client.count()
    print(f"📦 场景库加载完成，共 {total} 条场景")

    # ── Step 3: 读取数据行 ──
    items = []
    for r in range(2, ws.max_row + 1):
        t    = ws.cell(r, tc + 1).value or ""
        s    = ws.cell(r, sc + 1).value or ""
        d_raw = ws.cell(r, dc + 1).value or ""
        o    = ws.cell(r, oc + 1).value or ""
        c    = ws.cell(r, src + 1).value or ""
        z    = ws.cell(r, pl + 1).value or ""
        pain = ws.cell(r, pc + 1).value or ""
        if not t and not s:
            continue
        items.append((r, t, s, d_raw, o, c, z, pain))

    if not items:
        print("❌ 没有找到有效数据行")
        return

    # ── Step 3.5: 部门批量归一化 + 步骤详细度检查 ──
    raw_depts = [item[3] for item in items if item[3]]
    dept_mapping = normalize_depts_batch(raw_depts) if raw_depts else {}

    detailed_items, skipped = [], 0
    for item in items:
        rn, t, s, d_raw, o, c, z, pain = item
        ok_steps, short_label, full_reason = check_steps_detail(s)
        d = dept_mapping.get(d_raw, d_raw) if d_raw else ""
        if ok_steps:
            detailed_items.append((rn, t, s, d, d_raw, o, c, z, pain))
        else:
            print(f"  ⏭ 跳过第{rn}行：{full_reason}")
            skipped += 1

    if not detailed_items:
        print("❌ 所有条目步骤说明均不符合要求，请修改后重试")
        return

    print(f"📝 待匹配 {len(detailed_items)} 条有效需求（已过滤 {skipped} 条详细度不足）")

    # ── Step 4: 语义匹配 ──
    results = []
    for rn, t, s, d, d_raw, o, c, z, pain in detailed_items:
        q_text = f"{t} {s} {pain}"
        scene, sim = top_match(client, q_text, dept_std=d)
        results.append({
            "row": rn,         "title": t, "steps": s, "dept": d,
            "dept_raw": d_raw, "owner": o, "platform": z, "company": c,
            "pain": pain, "scene": scene, "sim": sim,
        })
        label = "已有场景" if sim >= HIGH_THRESHOLD else "待确认" if sim >= LOW_THRESHOLD else "新场景"
        color = "🔵" if sim >= HIGH_THRESHOLD else "🟡" if sim >= LOW_THRESHOLD else "🟢"
        if scene:
            print(f"  {color} 第{rn}行 [{d}] sim={sim:.3f} → {label} | 匹配：「{scene.get('scene_title','')}」")
        else:
            print(f"  🟢 第{rn}行 [{d}] → 新场景（场景库为空或无匹配）")

    # ── Step 5: 自动入库 ──
    imported = 0
    if args.auto_import:
        for res in results:
            if res["sim"] < LOW_THRESHOLD:
                sid = _new_id()
                import_scene(
                    client, sid, res["title"], res["steps"], res["dept"],
                    res["owner"], res["platform"], "", res["pain"],
                    dept_already_normalized=True,
                )
                imported += 1
        print(f"✅ 自动入库完成，{imported} 条新场景已入库")

    # ── Step 6: 生成结果 Excel ──
    out_path = input_path.replace(".xlsx", "_匹配结果.xlsx")
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "匹配结果"

    ws_out.row_dimensions[1].height = 30
    # AI 加的列：流程是否详细、Top1场景、Top1场景流程、相似度、自动判定、建议操作、是否入库
    col_headers = REQUIRED_HEADERS + [
        "流程是否详细", "Top1场景", "Top1场景流程",
        "相似度", "自动判定", "建议操作", "是否入库"
    ]

    # H-N 列表头统一青色（与 Top1场景流程 一致）
    _AI_HEADER_COLORS = [
        ("FF00B0F0", "FFFFFFFF"),  # H 流程是否详细
        ("FF00B0F0", "FFFFFFFF"),  # I Top1场景
        ("FF00B0F0", "FFFFFFFF"),  # J Top1场景流程
        ("FF00B0F0", "FFFFFFFF"),  # K 相似度
        ("FF00B0F0", "FFFFFFFF"),  # L 自动判定
        ("FF00B0F0", "FFFFFFFF"),  # M 建议操作
        ("FF00B0F0", "FFFFFFFF"),  # N 是否入库
    ]
    n_user_cols = len(REQUIRED_HEADERS)
    for c, h in enumerate(col_headers, 1):
        if c <= n_user_cols:
            fill, font = _STYLE_HEADER_FILL, _STYLE_HEADER_FONT
        else:
            ai_idx = c - n_user_cols - 1
            bg_color, fg_color = _AI_HEADER_COLORS[ai_idx]
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            font = Font(name="微软雅黑", bold=True, size=11, color=fg_color)
        _set_header(ws_out, 1, c, h, fill, font)

    # 部门列表头批注（挂在 A1 上）
    dept_cell = ws_out.cell(1, 1)
    try:
        from openpyxl.comments import Comment
        dept_cell.comment = Comment("此列已由AI进行部门归一化处理", "AI助手")
    except Exception:
        pass

    # A-G 原始 + H-N AI 列（共14列）
    _col_widths(ws_out, [14, 12, 22, 35, 30, 25, 15, 12, 20, 35, 12, 12, 15, 12])

    for res in results:
        sim = res["sim"]
        scene = res["scene"]
        label = "已有场景" if sim >= HIGH_THRESHOLD else "待确认" if sim >= LOW_THRESHOLD else "新场景"
        label_fill = (
            _STYLE_BLUE_FILL   if sim >= HIGH_THRESHOLD else
            _STYLE_YELLOW2_FILL if sim >= LOW_THRESHOLD else
            _STYLE_GREEN_FILL
        )

        # 建议操作映射
        action_map = {
            "已有场景": "不重复入库",
            "待确认":   "人工判断",
            "新场景":   "建议入库",
        }
        action = action_map.get(label, "")

        row_data = [
            res["dept"], res["owner"], res["title"], res["steps"],
            res["pain"], res["platform"], res.get("company", ""),
            "✓ 详细",
            scene.get("scene_title", "") if scene else "",
            scene.get("process_steps", "") if scene else "",
            f"{sim:.3f}",
            label,
            action,
            "请填写：是/否",
        ]
        for c, val in enumerate(row_data, 1):
            # 部门(A)用浅绿；H-N（AI 加的列）统一跟随判定结果颜色
            is_ai_col = c > n_user_cols
            if is_ai_col:
                fill = label_fill; font = None
                # N 列（是否入库）背景跟随，但字体保留提示样式
                if c == n_user_cols + 7:  # 是否入库
                    font = _STYLE_FEEDBACK_FONT
            else:
                fill = None; font = None
            align = "center" if c in (1,) or is_ai_col else "left"
            cell = _set_cell(ws_out, res["row"], c, val, fill, font, align=align)

    wb_out.save(out_path)
    print(f"\n✅ 匹配结果已保存：{out_path}")


if __name__ == "__main__":
    main()
